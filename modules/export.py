"""Export module – PDF and Excel export."""

import customtkinter as ctk
import os
from datetime import datetime
from pathlib import Path
from typing import Optional

from utils.database import Database
from utils.conversions import mm_to_inch, mm_to_moa, calculate_jump_thou
from utils.ballistics import es_color


class ExportFrame(ctk.CTkFrame):
    def __init__(self, parent, db: Database):
        super().__init__(parent, fg_color="transparent")
        self.db = db
        self._build_ui()

    def _build_ui(self):
        title = ctk.CTkLabel(
            self, text="Export", font=ctk.CTkFont(size=24, weight="bold"),
        )
        title.pack(pady=(10, 15), anchor="w", padx=20)

        content = ctk.CTkFrame(self, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=20)

        # Setup selection
        row = ctk.CTkFrame(content, fg_color="transparent")
        row.pack(fill="x", pady=10)
        ctk.CTkLabel(row, text="Setup :", width=120, anchor="w").pack(side="left")
        setups = self.db.get_setups()
        self._setups = {s["nom"]: s for s in setups}
        self.setup_combo = ctk.CTkComboBox(
            row, values=[s["nom"] for s in setups] if setups else ["Aucun"],
            width=300,
        )
        self.setup_combo.pack(side="left", padx=10)

        # Export directory
        row = ctk.CTkFrame(content, fg_color="transparent")
        row.pack(fill="x", pady=10)
        ctk.CTkLabel(row, text="Dossier :", width=120, anchor="w").pack(side="left")
        self.dir_entry = ctk.CTkEntry(row, width=350)
        self.dir_entry.pack(side="left", padx=10)
        self.dir_entry.insert(0, str(Path.home() / "Desktop"))
        ctk.CTkButton(row, text="Parcourir", width=80, command=self._browse).pack(side="left")

        # Export buttons
        btn_frame = ctk.CTkFrame(content, fg_color="transparent")
        btn_frame.pack(fill="x", pady=30)

        ctk.CTkButton(
            btn_frame, text="Export PDF", width=200, height=45,
            font=ctk.CTkFont(size=15, weight="bold"),
            command=self._export_pdf,
        ).pack(side="left", padx=10)

        ctk.CTkButton(
            btn_frame, text="Export Excel (.xlsx)", width=200, height=45,
            font=ctk.CTkFont(size=15, weight="bold"),
            fg_color="#2FA572", hover_color="#258A5C",
            command=self._export_excel,
        ).pack(side="left", padx=10)

        # Status
        self.status_label = ctk.CTkLabel(
            content, text="", font=ctk.CTkFont(size=14),
        )
        self.status_label.pack(pady=20)

    def _browse(self):
        from tkinter import filedialog
        folder = filedialog.askdirectory()
        if folder:
            self.dir_entry.delete(0, "end")
            self.dir_entry.insert(0, folder)

    def _get_filename(self, ext: str) -> str:
        setup_name = self.setup_combo.get().replace(" ", "_").replace("/", "-")
        calibre = ""
        setup = self._setups.get(self.setup_combo.get())
        if setup:
            calibre = setup.get("calibre", "").replace(" ", "").replace(".", "")
        date_str = datetime.now().strftime("%Y%m%d")
        return f"DevCharge_{setup_name}_{calibre}_{date_str}.{ext}"

    def _export_pdf(self):
        setup = self._setups.get(self.setup_combo.get())
        if not setup:
            self.status_label.configure(text="Sélectionnez un setup.", text_color="red")
            return

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import mm
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        except ImportError:
            self.status_label.configure(
                text="ReportLab non installé. pip install reportlab",
                text_color="red",
            )
            return

        output_dir = self.dir_entry.get()
        filename = self._get_filename("pdf")
        filepath = os.path.join(output_dir, filename)

        series_all = self.db.get_all_series(setup["id"])
        composants = self.db.get_composants(setup["id"])
        best = self.db.get_best_series(setup["id"])

        doc = SimpleDocTemplate(filepath, pagesize=A4,
                                leftMargin=20 * mm, rightMargin=20 * mm,
                                topMargin=15 * mm, bottomMargin=15 * mm)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle", parent=styles["Title"], fontSize=18, spaceAfter=10,
        )
        h2_style = ParagraphStyle(
            "H2", parent=styles["Heading2"], fontSize=13, spaceAfter=6, spaceBefore=12,
        )
        normal = styles["Normal"]

        elements = []

        # Title
        elements.append(Paragraph(f"Rapport de développement – {setup['nom']}", title_style))
        elements.append(Paragraph(
            f"Calibre : {setup.get('calibre', '')} | "
            f"Canon : {setup.get('longueur_canon_mm', '')} mm | "
            f"Twist : {setup.get('twist', '')}",
            normal,
        ))
        elements.append(Spacer(1, 8 * mm))

        # Components
        elements.append(Paragraph("Composants", h2_style))
        for comp in composants:
            elements.append(Paragraph(
                f"<b>{comp['type']}</b> : {comp.get('marque', '')} {comp.get('modele', '')}",
                normal,
            ))
        elements.append(Spacer(1, 5 * mm))

        # Series table
        elements.append(Paragraph("Séries enregistrées", h2_style))

        if series_all:
            data = [["Date", "Charge\n(gr)", "V moy\n(fps)", "ES\n(fps)",
                     "SD\n(fps)", "Grp\n(MOA)", "Jump\n(mm)"]]

            for s in series_all:
                d = s.get("distance_m", 100)
                grp_moa = mm_to_moa(s["groupement_mm"], d) if s.get("groupement_mm") else None

                data.append([
                    str(s.get("date", ""))[:10],
                    f'{s["charge_gr"]:.1f}',
                    f'{s["v_moy"]:.0f}' if s.get("v_moy") else "—",
                    f'{s["es"]:.1f}' if s.get("es") else "—",
                    f'{s["sd"]:.1f}' if s.get("sd") else "—",
                    f'{grp_moa:.2f}' if grp_moa else "—",
                    f'{s.get("jump_mm", 0):.2f}' if s.get("jump_mm") is not None else "—",
                ])

            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3B8ED0")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.HexColor("#F5F5F5"), colors.white]),
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph("Aucune série enregistrée.", normal))

        # Conclusions
        elements.append(Spacer(1, 8 * mm))
        elements.append(Paragraph("Conclusions", h2_style))
        if best.get("best_es") is not None:
            elements.append(Paragraph(
                f"Meilleur ES : <b>{best['best_es']:.1f} fps</b> "
                f"à {best.get('best_es_charge', '?'):.1f} gr",
                normal,
            ))
        if best.get("best_group_mm") is not None:
            moa = mm_to_moa(best["best_group_mm"], 100)
            elements.append(Paragraph(
                f"Meilleur groupement : <b>{best['best_group_mm']:.1f} mm "
                f"({moa:.2f} MOA)</b> à {best.get('best_group_charge', '?'):.1f} gr",
                normal,
            ))
        if best.get("charge_retenue") is not None:
            elements.append(Paragraph(
                f"Charge retenue : <b>{best['charge_retenue']:.1f} gr</b>",
                normal,
            ))

        doc.build(elements)
        self.status_label.configure(
            text=f"PDF exporté : {filepath}", text_color="#2FA572",
        )

    def _export_excel(self):
        setup = self._setups.get(self.setup_combo.get())
        if not setup:
            self.status_label.configure(text="Sélectionnez un setup.", text_color="red")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            self.status_label.configure(
                text="openpyxl non installé. pip install openpyxl",
                text_color="red",
            )
            return

        output_dir = self.dir_entry.get()
        filename = self._get_filename("xlsx")
        filepath = os.path.join(output_dir, filename)

        series_all = self.db.get_all_series(setup["id"])

        wb = Workbook()
        ws = wb.active
        ws.title = "Séries"

        # Header
        headers = [
            "Date", "Charge (gr)", "OAL (mm)", "CBTO (mm)", "Jump (mm)",
            "Nb coups", "V moy (fps)", "ES (fps)", "SD (fps)",
            "Groupement (mm)", "Groupement (MOA)", "Distance (m)",
            "Observations", "Charge retenue",
        ]
        header_fill = PatternFill(start_color="3B8ED0", end_color="3B8ED0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Data
        for row_idx, s in enumerate(series_all, 2):
            d = s.get("distance_m", 100)
            grp_moa = mm_to_moa(s["groupement_mm"], d) if s.get("groupement_mm") else None

            values = [
                str(s.get("date", ""))[:10],
                s["charge_gr"],
                s.get("oal_mm"),
                s.get("cbto_mm"),
                s.get("jump_mm"),
                s.get("nb_coups"),
                round(s["v_moy"], 1) if s.get("v_moy") else None,
                round(s["es"], 1) if s.get("es") else None,
                round(s["sd"], 1) if s.get("sd") else None,
                s.get("groupement_mm"),
                round(grp_moa, 2) if grp_moa else None,
                s.get("distance_m"),
                s.get("observations", ""),
                "Oui" if s.get("charge_retenue") else "",
            ]
            for col, val in enumerate(values, 1):
                ws.cell(row=row_idx, column=col, value=val)

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 25)

        # Vitesses sheet
        ws2 = wb.create_sheet("Vitesses")
        ws2.cell(row=1, column=1, value="Date").font = header_font
        ws2.cell(row=1, column=2, value="Charge (gr)").font = header_font
        max_v = max((len(s.get("vitesses", [])) for s in series_all), default=0)
        for i in range(max_v):
            ws2.cell(row=1, column=3 + i, value=f"V{i+1} (fps)").font = header_font

        for row_idx, s in enumerate(series_all, 2):
            ws2.cell(row=row_idx, column=1, value=str(s.get("date", ""))[:10])
            ws2.cell(row=row_idx, column=2, value=s["charge_gr"])
            for vi, v in enumerate(s.get("vitesses", [])):
                ws2.cell(row=row_idx, column=3 + vi, value=v)

        wb.save(filepath)
        self.status_label.configure(
            text=f"Excel exporté : {filepath}", text_color="#2FA572",
        )

    def refresh(self):
        setups = self.db.get_setups()
        self._setups = {s["nom"]: s for s in setups}
        self.setup_combo.configure(
            values=[s["nom"] for s in setups] if setups else ["Aucun"]
        )
